# Image handling imports
from skimage import data, io
from skimage.transform import rescale, resize
from skimage.util import img_as_ubyte

# Excel manipulation imports
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color
from openpyxl.utils.cell import get_column_letter

# argument parsing imports
import argparse

# parse inputs
parser = argparse.ArgumentParser(description='Transform an image into an Excel workbook with cells as pixels.')
parser.add_argument("--input_image", "-i", type=str, help="Path to image to transform, if not included will run"
                    " astronaut test image.")
parser.add_argument("--output_f", "-o", type=str, default="./ExcelImage.xlsx",
                    help="Path to output excel workbook.")
parser.add_argument("--dimensions", "-d", type=int, nargs=2, default=[-1, -1],
                    help="Optionally specify image dimensions in pixels in format --dimensions Xdimension Ydimension."
                         " Maximum dimensions are 350 x 350. If not specified image will retain current dimensions"
                         "or be truncated to obey maximum dimensions.")
args = parser.parse_args()

# constants
x_max = 350
y_max = 350
cell_pixel_dim = 12

# Sort of lazy, but on my monitor at 100% zoom Excel is showing 97*12 pixels vertical, 220*12 horizontal
default_x_zoom = 97 * 12
default_y_zoom = 220 * 12

# input image
if args.input_image:
    # read in input image
    input_image = io.imread(args.input_image)
else:
    input_image = data.astronaut()

# if dimensions are specified, resize image
new_x, new_y = args.dimensions[0], args.dimensions[1]
if new_x != -1 and new_y != -1:
    input_image = resize(input_image, [new_x, new_y], anti_aliasing=True)

# Get X/Y dimensions of image
l_x, l_y = input_image.shape[0], input_image.shape[1]

# If X or Y above max, scale image down maintaining ratio

# x scale factor
if l_x > x_max:
    x_sf = x_max / l_x
else:
    x_sf = 1

# y scale factor
if l_y > y_max:
    y_sf = y_max / l_y
else:
    y_sf = 1

# determine worst case scale factor
if x_sf < y_sf:
    sf = x_sf
else:
    sf = y_sf

# Scale image
if sf < 1:
    input_image = img_as_ubyte(rescale(input_image, sf, anti_aliasing=True))
else:
    input_image = img_as_ubyte(input_image)

# get new x/y size
l_x, l_y = input_image.shape[0], input_image.shape[1]

# Convert image matrix to excel matrix
wb = Workbook()
ws = wb.active
ws.title = "converted_image"

set_col_height = False

# Output excel workbook containing cell pixelated image
for row in range(0, l_x):

    ws.row_dimensions[row+1].height = 4.5

    for col in range(0, l_y):

        if not set_col_height:
            ws.column_dimensions[get_column_letter(col+1)].width = 0.83

        # Determine RGB from image array, include opacity if it is in image
        if input_image.shape[2] > 3:
            cell_hex = "{:02X}".format(input_image[row, col, 3]) + "{:02X}".format(input_image[row, col, 0]) + \
                       "{:02X}".format(input_image[row, col, 1]) + "{:02X}".format(input_image[row, col, 2])
        else:
            cell_hex = "{:02X}".format(input_image[row, col, 0]) + "{:02X}".format(input_image[row, col, 1]) \
                       + "{:02X}".format(input_image[row, col, 2])

        # Set color using styles, Color takes ARGB hex input as AARRGGBB
        cell_color = Color(cell_hex)

        # Set cell fill
        sel_cell = ws.cell(column=col+1, row=row+1) # , value=1
        sel_cell.fill = PatternFill("solid", fgColor=cell_color)

    set_col_height = True

# Add a " " to a cell at the bottom right corner to enable zoom to fit
sel_cell = ws.cell(column=l_y+1, row=l_x+1, value=1)

# Set zoom scale according to dimensions of photo
pixels_x = l_x * cell_pixel_dim
pixels_y = l_y * cell_pixel_dim

zoom_scale_x = int((default_x_zoom / pixels_x) * 100)
zoom_scale_y = int((default_y_zoom / pixels_y) * 100)

if zoom_scale_x < zoom_scale_y:
    ws.sheet_view.zoomScale = zoom_scale_x
else:
    ws.sheet_view.zoomScale = zoom_scale_y

# Output workbook
wb.save(args.output_f)
